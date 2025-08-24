# pages/04_Rolagem.py — Rolagem + Capacidade com auditoria UC vs Sellout
from __future__ import annotations
import io
import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridUpdateMode
from st_aggrid.shared import JsCode

from core.rolagem_estoque import (
    build_tabela_inicial_with_audit,
    load_vol_prod_ytg,
    load_vendas_res,           # baseline (res_working) LONG
    MONTH_MAP_PT,
)

st.set_page_config(page_title="Rolagem de Estoque", page_icon="📦", layout="wide")

# =============================== Helpers UI ================================= #
def _fmt_int(scale: int) -> JsCode:
    return JsCode(
        f"""
        function(params){{
          if (params.value === null || params.value === undefined) return '-';
          var v = Number(params.value) / {scale};
          v = Math.round(v);
          return v.toString().replace(/\\B(?=(\\d{{3}})+(?!\\d))/g, '.');
        }}
        """
    )

def _row_style_total() -> JsCode:
    return JsCode(
        """
        function(params) {
          if (params.node && params.node.rowPinned === 'top') {
            return {backgroundColor: '#FFF3B0', fontWeight: 'bold'};
          }
          return null;
        }
        """
    )

def _cell_style_estoque() -> JsCode:
    return JsCode(
        """
        function(params) {
          const hn = params.colDef.headerName || params.colDef.field || "";
          if (hn.includes("Estoque")) {
            let base = {backgroundColor: "#f0f0f0", fontWeight: "bold"};
            const val = Number(params.value);
            if (!isNaN(val)) {
              if (val > 0) base.color = "#0a7f27";
              else if (val < 0) base.color = "#b00020";
            }
            return base;
          }
          return {};
        }
        """
    )

def _cell_style_delta() -> JsCode:
    return JsCode(
        """
        function(params) {
          const val = Number(params.value);
          if (!isNaN(val)) {
            if (val >= 0) return {fontWeight: "bold", color: "#0a7f27"};
            else return {fontWeight: "bold", color: "#b00020"};
          }
          return {};
        }
        """
    )

# ------------------ ColDefs (headers 2 níveis) ------------------------------ #
def _build_grouped_coldefs_tab2(meses: list[int], scale: int):
    fmt = _fmt_int(scale); style_ef = _cell_style_estoque()
    W_NUM = 96  # ~10 caracteres

    coldefs = [
        {
            "field": "Família Comercial", "headerName": "Família Comercial",
            "editable": False, "resizable": True, "pinned": "left",
            "width": 220, "minWidth": 180
        },
        {
            "field": "Inicial", "headerName": "Inicial",
            "type": "numericColumn", "cellClass": "ag-right-aligned-cell",
            "valueFormatter": fmt, "editable": False, "resizable": True,
            "width": W_NUM, "minWidth": W_NUM
        },
    ]
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        coldefs.append({
            "headerName": mn, "marryChildren": True,
            "children": [
                {"field": f"{mn}__prod", "headerName": "Produção", "type": "numericColumn",
                 "cellClass": "ag-right-aligned-cell", "valueFormatter": fmt, "editable": False,
                 "resizable": True, "width": W_NUM, "minWidth": W_NUM},
                {"field": f"{mn}__vend", "headerName": "Vendas", "type": "numericColumn",
                 "cellClass": "ag-right-aligned-cell", "valueFormatter": fmt, "editable": False,
                 "resizable": True, "width": W_NUM, "minWidth": W_NUM},
                {"field": f"{mn}__ef", "headerName": "Estoque", "type": "numericColumn",
                 "cellClass": "ag-right-aligned-cell", "valueFormatter": fmt, "cellStyle": style_ef,
                 "editable": False, "resizable": True, "width": W_NUM, "minWidth": W_NUM},
            ]
        })
    return coldefs

def _build_grouped_coldefs_tab1(meses: list[int], scale: int):
    fmt = _fmt_int(scale); style_delta = _cell_style_delta()
    W_NUM = 96
    coldefs = [
        {"field": "Tecnologia", "headerName": "Tecnologia",
         "editable": False, "resizable": True, "pinned": "left", "width": 200, "minWidth": 160},
        {"field": "Capacidade", "headerName": "Capacidade",
         "type": "numericColumn", "cellClass": "ag-right-aligned-cell",
         "valueFormatter": fmt, "editable": False, "resizable": True, "width": W_NUM, "minWidth": W_NUM},
    ]
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        coldefs.append({
            "headerName": mn, "marryChildren": True,
            "children": [
                {"field": f"{mn}__nec", "headerName": "Necessidade", "type": "numericColumn",
                 "cellClass": "ag-right-aligned-cell", "valueFormatter": fmt, "editable": False,
                 "resizable": True, "width": W_NUM, "minWidth": W_NUM},
                {"field": f"{mn}__delta", "headerName": "Delta", "type": "numericColumn",
                 "cellClass": "ag-right-aligned-cell", "valueFormatter": fmt, "cellStyle": style_delta,
                 "editable": False, "resizable": True, "width": W_NUM, "minWidth": W_NUM},
            ]
        })
    return coldefs

def _grid_grouped(df: pd.DataFrame, coldefs, key: str, pinned_top_row_data=None):
    """Sem scroll vertical (autoHeight). Scroll horizontal aparece quando necessário, estável."""
    if df is None or df.empty:
        st.warning("Sem dados para exibir.", icon="⚠️")
        return

    opts = {
        "columnDefs": coldefs,
        "defaultColDef": {
            "sortable": True, "filter": False, "resizable": True,
            "width": 96, "minWidth": 96  # fallback p/ 10 chars
        },
        "ensureDomOrder": True,
        "suppressMovableColumns": True,
        "rowHeight": 34,
        "animateRows": False,
        "domLayout": "autoHeight",        # SEM scroll vertical
        "suppressSizeToFit": True,        # NÃO tentar caber
        "suppressColumnVirtualisation": True,  # evita recalcular largura
        "alwaysShowHorizontalScroll": True,    # força barra horizontal
        "getRowStyle": _row_style_total(),
        "maintainColumnOrder": True,
        "suppressAnimationFrame": True,
    }
    if pinned_top_row_data:
        opts["pinnedTopRowData"] = pinned_top_row_data

    AgGrid(
        df,
        gridOptions=opts,
        data_return_mode="AS_INPUT",
        update_mode=GridUpdateMode.NO_UPDATE,
        allow_unsafe_jscode=True,
        theme="balham",
        fit_columns_on_grid_load=False,   # NÃO ajustar para caber
        key=key,
    )

# ============================ Normalização Vendas ============================ #
_UI_MONTH_MAP = {
    "jan":1,"janeiro":1,"jan.":1, "fev":2,"fevereiro":2,"fev.":2,
    "mar":3,"março":3,"marco":3,"mar.":3, "abr":4,"abril":4,"abr.":4,
    "mai":5,"maio":5,"mai.":5, "jun":6,"junho":6,"jun.":6,
    "jul":7,"julho":7,"jul.":7, "ago":8,"agosto":8,"ago.":8,
    "set":9,"setembro":9,"set.":9,"sep":9, "out":10,"outubro":10,"out.":10,"oct":10,
    "nov":11,"novembro":11,"nov.":11, "dez":12,"dezembro":12,"dez.":12,"dec":12,
}
_UI_MONTH_MAP.update({k.capitalize(): v for k, v in _UI_MONTH_MAP.items()})

def _normalize_ui_wide_to_long(df_wide: pd.DataFrame, ano: int) -> pd.DataFrame:
    if df_wide is None or len(df_wide) == 0:
        return pd.DataFrame(columns=["Família Comercial","ano","mes","vendas"])
    df = df_wide.copy()
    if "Família Comercial" not in df.columns:
        df = df.reset_index().rename(columns={df.columns[0]:"Família Comercial"})
    rows = []
    inv_month_pt = {v:k for k, v in MONTH_MAP_PT.items()}
    for col in df.columns:
        if col == "Família Comercial":
            continue
        key = str(col).strip()
        mes = _UI_MONTH_MAP.get(key.lower()) or inv_month_pt.get(key)
        if mes is None:
            continue
        tmp = df[["Família Comercial", col]].copy()
        tmp["ano"] = int(ano); tmp["mes"] = int(mes)
        tmp = tmp.rename(columns={col:"vendas"})
        rows.append(tmp)
    if not rows:
        return pd.DataFrame(columns=["Família Comercial","ano","mes","vendas"])
    out = pd.concat(rows, ignore_index=True)
    out["vendas"] = pd.to_numeric(out["vendas"], errors="coerce").fillna(0.0)
    return out[["Família Comercial","ano","mes","vendas"]]

# =============================== Bootstrap + UC ============================== #
def bootstrap_ui_volumes_from_res(year: int, meses: list[int]) -> None:
    """Executa 1x se NÃO houver nenhum dado de UI; não sobrescreve UI existente."""
    if "volumes_wide" in st.session_state:
        wide = st.session_state["volumes_wide"]
        if isinstance(wide, dict) and year in wide and not pd.DataFrame(wide[year]).empty:
            return
    if "volumes_edit" in st.session_state and not pd.DataFrame(st.session_state["volumes_edit"]).empty:
        return

    vend_res = load_vendas_res(year, meses)  # LONG baseline
    if vend_res.empty:
        return

    df = vend_res.copy()
    # tentar filtrar UC se existir tipo/métrica
    type_cols = [c for c in df.columns if str(c).lower().strip() in {"metric","tipo","tipo_volume","medida","tipo de volume"}]
    if type_cols:
        tcol = type_cols[0]
        df[tcol] = df[tcol].astype(str).str.lower().str.strip()
        uc_keys = {"uc","volume_uc","volume uc","vol_uc","unidades","unit","unidade"}
        df = df[df[tcol].isin(uc_keys)] or df

    if "vendas" not in df.columns:
        num_cols = [c for c in df.columns if c not in {"Família Comercial","ano","mes"} and pd.api.types.is_numeric_dtype(df[c])]
        uc_num = [c for c in num_cols if "uc" in str(c).lower()]
        if uc_num: df = df.rename(columns={uc_num[0]: "vendas"})
        elif num_cols: df = df.rename(columns={num_cols[0]: "vendas"})

    keep = [c for c in ["Família Comercial","ano","mes","vendas"] if c in df.columns]
    df = df[keep].copy()
    df["ano"] = pd.to_numeric(df["ano"], errors="coerce").astype("Int64")
    df["mes"] = pd.to_numeric(df["mes"], errors="coerce").astype("Int64")
    df["vendas"] = pd.to_numeric(df["vendas"], errors="coerce").fillna(0.0)
    df = df.groupby(["Família Comercial","ano","mes"], as_index=False)["vendas"].sum()

    st.session_state["volumes_edit"] = df  # LONG para edição na UI

def _filter_to_volume_uc_long(df: pd.DataFrame, ano: int) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["Família Comercial","ano","mes","vendas"])
    out = df.copy()

    # normaliza nome da família
    fam_col = next((c for c in out.columns if str(c).lower().strip() in
                   {"família comercial","familia comercial","familia","familia_comercial"}), "Família Comercial")
    if fam_col != "Família Comercial":
        out = out.rename(columns={fam_col: "Família Comercial"})

    # reforço: manter apenas UC quando houver coluna de tipo/metric
    tcol = None
    for c in out.columns:
        lc = str(c).lower().strip()
        if lc in {"metric","tipo","tipo_volume","medida","tipo de volume"}:
            tcol = c; break
    if tcol:
        out[tcol] = out[tcol].astype(str).str.lower().str.strip()
        uc_keys = {"uc","volume_uc","volume uc","vol_uc","unidades","unit","unidade"}
        out = out[out[tcol].isin(uc_keys)]

    # se não há coluna vendas, pega coluna numérica mais provável (que contenha uc)
    if "vendas" not in out.columns:
        num_cols = [c for c in out.columns if c not in {"Família Comercial","ano","mes"} and pd.api.types.is_numeric_dtype(out[c])]
        uc_num = [c for c in num_cols if "uc" in str(c).lower()]
        if uc_num: out = out.rename(columns={uc_num[0]: "vendas"})
        elif num_cols: out = out.rename(columns={num_cols[0]: "vendas"})

    out["ano"] = pd.to_numeric(out["ano"], errors="coerce").astype("Int64")
    out["mes"] = pd.to_numeric(out["mes"], errors="coerce").astype("Int64")
    out["vendas"] = pd.to_numeric(out["vendas"], errors="coerce").fillna(0.0)
    out = out[out["ano"] == int(ano)]
    return out[["Família Comercial","ano","mes","vendas"]].groupby(
        ["Família Comercial","ano","mes"], as_index=False
    )["vendas"].sum()

def _load_vendas_from_toggle(ano: int, meses: list[int], use_ui: bool):
    """
    Toggle ON  -> preferir UI: volumes_wide[ano] -> volumes_edit; se ambos vazios, bootstrap 1x.
    Toggle OFF -> baseline (res_working) SEM bootstrap.
    """
    source = "none"
    vendas_debug_pre = None

    if use_ui:
        # 1) volumes_wide[ano] (UI de planilha wide)
        if "volumes_wide" in st.session_state:
            wide = st.session_state["volumes_wide"]
            if isinstance(wide, dict) and ano in wide:
                dfw = pd.DataFrame(wide[ano])
                if not dfw.empty:
                    vend_long = _normalize_ui_wide_to_long(dfw, ano)
                    vendas_debug_pre = vend_long.copy()
                    vend_long = _filter_to_volume_uc_long(vend_long, ano)
                    vend_long = vend_long[vend_long["mes"].isin(meses)]
                    source = "UI (wide)"
                    return vend_long, source, vendas_debug_pre

        # 2) volumes_edit (UI long)
        if "volumes_edit" in st.session_state:
            ve = pd.DataFrame(st.session_state["volumes_edit"])
            if not ve.empty:
                vendas_debug_pre = ve.copy()
                ve = _filter_to_volume_uc_long(ve, ano)
                ve = ve[ve["mes"].isin(meses)]
                source = "UI (edit)"
                return ve, source, vendas_debug_pre

        # 3) Bootstrap 1x para liberar edição
        bootstrap_ui_volumes_from_res(ano, meses)
        if "volumes_edit" in st.session_state:
            ve = pd.DataFrame(st.session_state["volumes_edit"])
            vendas_debug_pre = ve.copy()
            ve = _filter_to_volume_uc_long(ve, ano)
            ve = ve[ve["mes"].isin(meses)]
            source = "Bootstrap (baseline→UI)"
            return ve, source, vendas_debug_pre

        return pd.DataFrame(columns=["Família Comercial","ano","mes","vendas"]), source, vendas_debug_pre
    else:
        vend_res = load_vendas_res(ano, meses)
        vendas_debug_pre = vend_res.copy()
        out = _filter_to_volume_uc_long(vend_res, ano)
        out = out[out["mes"].isin(meses)]
        source = "Baseline (res_working)"
        return out, source, vendas_debug_pre

# ============================ Título + Sidebar ============================== #
left, right = st.columns([1, 0.6])
with left:
    st.markdown("## Rolagem de Estoque — YTG (Produção YTG + Vendas da UI + Estoque)")
with st.sidebar:
    st.subheader("Controles")
    escala_label = st.selectbox("Escala", ["1x","1.000x","1.000.000x"], index=0, key="escala_rolagem")
    scale = {"1x":1, "1.000x":1000, "1.000.000x":1_000_000}[escala_label]
    use_ui = st.toggle(
        "Usar Vendas da UI (com bootstrap automático)",
        value=True, key="toggle_vendas_rolagem",
        help="ON: Vendas da UI; se a UI ainda não populou, carregamos baseline 1x. OFF: baseline fixo."
    )
    if st.button("Recalcular", use_container_width=True, key="btn_recalc_rolagem"):
        st.session_state["rolagem_bump_ini"] = st.session_state.get("rolagem_bump_ini", 0) + 1
        st.rerun()

# ============================ Cálculo da Rolagem ============================= #
df_ini_long, _audit_ini = build_tabela_inicial_with_audit()
if df_ini_long.empty:
    st.warning("Sem dados para exibir.", icon="⚠️"); st.stop()

# Calendário a partir do YTG
ytg = load_vol_prod_ytg()
if ytg.empty:
    st.error("Produção YTG não encontrada."); st.stop()
ano = int(ytg["ano"].dropna().astype(int).max())
meses = sorted(ytg[ytg["ano"] == ano]["mes"].dropna().astype(int).unique().tolist())
if not meses:
    st.error("Calendário YTG sem meses válidos."); st.stop()
first_month = int(min(meses))

# Inicial (único no primeiro mês)
inicial = (
    df_ini_long[df_ini_long["mes"] == first_month]
    .groupby("Família Comercial", as_index=False)["estoque_inicial"]
    .sum()
    .rename(columns={"estoque_inicial": "Inicial"})
)

# Produção (YTG do ano)
ytg_ano = ytg[ytg["ano"] == ano]
pivot_prod = (ytg_ano.pivot_table(index="Família Comercial", columns="mes",
                                  values="producao", aggfunc="sum", fill_value=0.0)
              .reset_index())

# Vendas conforme TOGGLE
vendas_long, vendas_source, vendas_debug_pre = _load_vendas_from_toggle(ano, meses, use_ui=use_ui)
pivot_vend = pd.DataFrame()
if not vendas_long.empty:
    pivot_vend = (vendas_long
                  .pivot_table(index="Família Comercial", columns="mes",
                               values="vendas", aggfunc="sum", fill_value=0.0)
                  .reset_index())

# Montagem base (Rolagem)
base = inicial.copy()
for m in meses:
    mn = MONTH_MAP_PT.get(int(m), str(m))
    col_p = f"{mn}__prod"
    col_v = f"{mn}__vend"
    col_e = f"{mn}__ef"

    if m in pivot_prod.columns:
        base = base.merge(
            pivot_prod[["Família Comercial", m]].rename(columns={m: col_p}),
            on="Família Comercial", how="left"
        )
    else:
        base[col_p] = 0.0

    if not pivot_vend.empty and m in pivot_vend.columns:
        base = base.merge(
            pivot_vend[["Família Comercial", m]].rename(columns={m: col_v}),
            on="Família Comercial", how="left"
        )
    else:
        base[col_v] = 0.0

    base[col_e] = 0.0  # preenchido abaixo

# Conversão + Estoque Final (pode negativo)
for c in [c for c in base.columns if c != "Família Comercial"]:
    base[c] = pd.to_numeric(base[c], errors="coerce").fillna(0.0)

for idx, row in base.iterrows():
    saldo = float(row["Inicial"])
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        ef = saldo + float(row[f"{mn}__prod"]) - float(row[f"{mn}__vend"])
        base.at[idx, f"{mn}__ef"] = ef
        saldo = ef

# ======================== Capacidade Produtiva (Tab 1) ====================== #
def _load_tecnologias_familia() -> pd.DataFrame:
    paths = [
        "data/tec_poh/tecnologias_familia.xlsx",   # caminho informado
        "data/tecnologia/tecnologias_familia.xlsx" # fallback antigo
    ]
    for path in paths:
        try:
            df = pd.read_excel(path)
            fam_col = next((c for c in df.columns if str(c).lower().strip() in
                           {"família comercial","familia comercial","familia","familia_comercial"}), None)
            tec_col = next((c for c in df.columns if str(c).lower().strip() in
                           {"tecnologia","technology","tec"}), None)
            if not fam_col or not tec_col:
                continue
            df = df.rename(columns={fam_col:"Família Comercial", tec_col:"Tecnologia"})
            df["Família Comercial"] = df["Família Comercial"].astype(str).str.strip()
            df["Tecnologia"] = df["Tecnologia"].astype(str).str.strip()
            return df[["Família Comercial","Tecnologia"]].dropna()
        except Exception:
            continue
    return pd.DataFrame(columns=["Família Comercial","Tecnologia"])

def _load_capacidade_tecnologias() -> pd.DataFrame:
    path = "data/cap_poh/capacidade_tecnologias.xlsx"
    try:
        df = pd.read_excel(path)
    except Exception:
        return pd.DataFrame(columns=["Tecnologia","Capacidade"])
    tec_col = next((c for c in df.columns if str(c).lower().strip() in
                   {"tecnologia","technology","tec"}), None)
    cap_col = next((c for c in df.columns if str(c).lower().strip() in
                   {"capacidade","capacity","cap"}), None)
    if not tec_col or not cap_col:
        return pd.DataFrame(columns=["Tecnologia","Capacidade"])
    df = df.rename(columns={tec_col:"Tecnologia", cap_col:"Capacidade"})
    df["Tecnologia"] = df["Tecnologia"].astype(str).str.strip()
    # converte pt-BR "6.500.000" → 6500000
    df["Capacidade"] = (df["Capacidade"].astype(str)
                        .str.replace(".", "", regex=False)
                        .str.replace(",", ".", regex=False))
    df["Capacidade"] = pd.to_numeric(df["Capacidade"], errors="coerce").fillna(0.0)
    return df[["Tecnologia","Capacidade"]].dropna()

df_map = _load_tecnologias_familia()
df_cap = _load_capacidade_tecnologias()

# Necessidade por TECNOLOGIA:
#   Necessidade_mês = VENDAS(UI/baseline) + |negativo da rolagem (estoque final < 0, antes do clamp)|
# Aqui usamos apenas as VENDAS (já calculadas) e o saldo negativo implícito
# → como o EF já está calculado mês a mês, adotamos necessidade como as VENDAS
#   + (consumo extra por estoques negativos). Para manter simples: necessidade = VENDAS (UI/baseline).
#   Se quiser somar negativos da rolagem, poderíamos recomputar EF "sem clamp" e adicionar o |negativo|.
#   Mantemos VENDAS como proxy principal conforme solicitado.
# Constrói long de vendas por família/mês:
vend_long_for_tec = []
for m in meses:
    mn = MONTH_MAP_PT[m]
    if f"{mn}__vend" in base.columns:
        tmp = base[["Família Comercial", f"{mn}__vend"]].copy()
        tmp = tmp.rename(columns={f"{mn}__vend": "vendas"})
        tmp["mes"] = int(m)
        vend_long_for_tec.append(tmp)
if vend_long_for_tec:
    vend_long_for_tec = pd.concat(vend_long_for_tec, ignore_index=True)
else:
    vend_long_for_tec = pd.DataFrame(columns=["Família Comercial","vendas","mes"])

# família → tecnologia
tec_need = (vend_long_for_tec
            .merge(df_map, on="Família Comercial", how="left")
            .dropna(subset=["Tecnologia"]))

tec_need = (tec_need.groupby(["Tecnologia","mes"], as_index=False)["vendas"]
            .sum().rename(columns={"vendas":"necessidade"}))


# monta tabela de capacidade × meses
cap_wide = pd.DataFrame(columns=["Tecnologia","Capacidade"])
if not df_cap.empty:
    cap_wide = df_cap.copy()
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        col_nec, col_del = f"{mn}__nec", f"{mn}__delta"
        need_m = tec_need[tec_need["mes"] == int(m)][["Tecnologia","necessidade"]]
        cap_wide = cap_wide.merge(need_m.rename(columns={"necessidade": col_nec}),
                                  on="Tecnologia", how="left")
        cap_wide[col_nec] = pd.to_numeric(cap_wide[col_nec], errors="coerce").fillna(0.0)
        cap_wide[col_del] = cap_wide["Capacidade"] - cap_wide[col_nec]

# ========================== Exportação (autofit leve) ======================= #
def _calc_width_chars(values, head1: str, head2: str, is_numeric: bool) -> int:
    def fmt_br_int(x):
        try:
            n = int(round(float(str(x).replace('.', '').replace(',', '.'))))
        except Exception:
            return ""
        return f"{n:,}".replace(",", ".")
    h1 = len(str(head1 or "")); h2 = len(str(head2 or ""))
    max_len = max(h1, h2)
    if is_numeric:
        for v in values: max_len = max(max_len, len(fmt_br_int(v)))
        min_w, max_w = 6, 18
    else:
        for v in values: max_len = max(max_len, len("" if v is None else str(v)))
        min_w, max_w = 6, 24
    return max(min_w, min(max_len + 1, max_w))

def _export_capacidade_bytes(df: pd.DataFrame, meses: list[int]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    header_top = ["Tecnologia", "Capacidade"]; header_bot = ["", ""]
    export_cols = ["Tecnologia", "Capacidade"]
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        header_top += [mn, ""]; header_bot += ["Necessidade", "Delta"]; export_cols += [f"{mn}__nec", f"{mn}__delta"]

    wb = Workbook(); ws = wb.active; ws.title = "Capacidade"
    ws.append(header_top); ws.append(header_bot)

    col_idx = 3
    for _m in meses:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
        col_idx += 2

    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
    bold = Font(bold=True); center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_col = 2 + 2*len(meses)
    for r in (1, 2):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c); cell.font = bold; cell.alignment = center; cell.fill = header_fill

    df_out = df.copy()
    for col in export_cols:
        if col not in df_out.columns:
            df_out[col] = []
    df_out = df_out[export_cols]

    for _, row in df_out.iterrows():
        ws.append([row[c] for c in export_cols])

    num_fmt = "#,##0"
    for j, col_name in enumerate(export_cols, start=1):
        is_num = (col_name != "Tecnologia")
        if is_num:
            for r in range(3, ws.max_row+1):
                ws.cell(row=r, column=j).number_format = num_fmt
        col_vals = [ws.cell(row=r, column=j).value for r in range(3, ws.max_row+1)]
        w = _calc_width_chars(col_vals, header_top[j-1], header_bot[j-1], is_num)
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "C3"
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()

def _export_rolagem_bytes(df: pd.DataFrame, meses: list[int]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill

    header_top = ["Família Comercial", "Inicial"]; header_bot = ["", ""]
    export_cols = ["Família Comercial", "Inicial"]
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        header_top += [mn, "", ""]; header_bot += ["Produção", "Vendas", "Estoque"]; export_cols += [f"{mn}__prod", f"{mn}__vend", f"{mn}__ef"]

    wb = Workbook(); ws = wb.active; ws.title = "Rolagem"
    ws.append(header_top); ws.append(header_bot)

    col_idx = 3
    for _m in meses:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
        col_idx += 3

    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
    bold = Font(bold=True); center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_col = 2 + 3*len(meses)
    for r in (1, 2):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c); cell.font = bold; cell.alignment = center; cell.fill = header_fill

    df_out = df.copy()
    for col in export_cols:
        if col not in df_out.columns:
            df_out[col] = []
    df_out = df_out[export_cols]

    for _, row in df_out.iterrows():
        ws.append([row[c] for c in export_cols])

    num_fmt = "#,##0"
    for j, col_name in enumerate(export_cols, start=1):
        is_num = (col_name != "Família Comercial")
        if is_num:
            for r in range(3, ws.max_row+1):
                ws.cell(row=r, column=j).number_format = num_fmt
        col_vals = [ws.cell(row=r, column=j).value for r in range(3, ws.max_row+1)]
        w = _calc_width_chars(col_vals, header_top[j-1], header_bot[j-1], is_num)
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "C3"
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()

# =============================== Botões no TOPO ============================= #
with right:
    cap_bytes = _export_capacidade_bytes(cap_wide, meses)
    roll_bytes = _export_rolagem_bytes(base, meses)
    b1, b2 = st.columns(2)
    with b1:
        st.download_button("Exportar Capacidade", cap_bytes, file_name=f"Capacidade_{ano}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="btn_export_capacidade")
    with b2:
        st.download_button("Exportar Rolagem", roll_bytes, file_name=f"Rolagem_Estoque_{ano}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="btn_export_rolagem")

# ================================ Abas UI =================================== #
tab1, tab2 = st.tabs(["Capacidade Produtiva", "Rolagem de Estoque"])

with tab1:
    st.markdown("#### Capacidade Produtiva — Tecnologia × Mês")
    coldefs_t1 = _build_grouped_coldefs_tab1(meses, scale=scale)
    _grid_grouped(cap_wide, coldefs_t1, key=f"grid_capacity_autoheight_{scale}_{use_ui}")

with tab2:
    st.markdown("#### Rolagem — Inicial | Mês × (Produção, Vendas, Estoque)")
    coldefs_t2 = _build_grouped_coldefs_tab2(meses, scale=scale)
    # TOTAL GERAL (linha fixa no topo)
    total_row = {"Família Comercial": "TOTAL GERAL"}
    total_row["Inicial"] = float(base["Inicial"].sum()) if "Inicial" in base.columns else 0.0
    for m in meses:
        mn = MONTH_MAP_PT.get(int(m), str(m))
        p_col, v_col, e_col = f"{mn}__prod", f"{mn}__vend", f"{mn}__ef"
        if p_col in base.columns: total_row[p_col] = float(base[p_col].sum())
        if v_col in base.columns: total_row[v_col] = float(base[v_col].sum())
        if e_col in base.columns: total_row[e_col] = float(base[e_col].sum())

    _grid_grouped(
        base,
        coldefs_t2,
        key=f"grid_roll_grouped_autoheight_{scale}_{use_ui}",
        pinned_top_row_data=[total_row],
    )

# ============================== Diagnóstico ================================= #
with st.expander("🔎 Diagnóstico de Fontes (para auditoria)", expanded=False):
    st.write(f"**Ano calendário (YTG):** {ano}")
    st.write(f"**Meses:** {meses}")
    st.write(f"**Fonte de VENDAS (toggle={'ON' if use_ui else 'OFF'}):** {vendas_source}")

    # Auditoria de duplicação UC vs Sellout
    if vendas_debug_pre is not None and not pd.DataFrame(vendas_debug_pre).empty:
        vd = vendas_debug_pre.copy()
        st.markdown("**Vendas (pré-filtro UC) — amostra 5:**")
        st.dataframe(vd.head(5))

        # tenta descobrir a coluna de tipo
        tcol = None
        for c in vd.columns:
            lc = str(c).lower().strip()
            if lc in {"metric","tipo","tipo_volume","medida","tipo de volume"}:
                tcol = c; break
        if tcol:
            tmp = vd.copy()
            tmp[tcol] = tmp[tcol].astype(str).str.lower().str.strip()
            if "vendas" not in tmp.columns:
                # escolhe col numérica para somar
                num_cols = [c for c in tmp.columns if c not in {"Família Comercial","ano","mes"} and pd.api.types.is_numeric_dtype(tmp[c])]
                if num_cols:
                    tmp = tmp.rename(columns={num_cols[0]:"vendas"})
            if "vendas" in tmp.columns:
                soma_tipo = (tmp.groupby([tcol,"mes"], dropna=False)["vendas"]
                               .sum().reset_index().sort_values(["mes",tcol]))
                st.markdown("**Soma por tipo (para checar UC vs Sellout):**")
                st.dataframe(soma_tipo)
        st.write(f"**Shape pré-filtro:** {vd.shape}")

    if not vendas_long.empty:
        st.markdown("**Vendas (pós-filtro UC) — amostra 5:**")
        st.dataframe(vendas_long.head(5))
        st.write(f"**Shape pós-filtro:** {vendas_long.shape}")

    # Capacidade base
    st.markdown("**Capacidade base (cap_poh) — primeiras linhas:**")
    st.dataframe(_load_capacidade_tecnologias().head(10))
